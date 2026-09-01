"""
ingest.py — deterministic converter: content-plan spreadsheet -> Publer CSVs.

Reads the hand-filled `.ods`/`.xls`/`.xlsx` content plan (sheet "Контент-план") and
slices it into one `PlanCsv` per content unit — post, clip, reel, video,
shorts. Each unit reads a fixed set of ABSOLUTE spreadsheet columns declared
in `PLAN_SPECS`; units that share a date (e.g. FB posts and FB clips) each
have their own time column.

A row becomes a post as soon as the unit has been started there — any of its
text, media or title filled in. Whatever is filled goes to the CSV, gaps and
all: a missing time may well be deliberate, to be scheduled inside Publer.
The gaps come back as warnings. A unit nobody started anywhere produces no
file at all.

This is the front of the pipeline's ingestion stage. It is NOT an agent: the
mapping from spreadsheet columns to CSV columns is a fixed table with a
single correct answer.

Output files are named `{plan file name}_{unit suffix}.csv`. Rows are
`CsvRow` objects with the target `Platform` already assigned, ready to feed
straight into the orchestrator. When written to disk they use the 12-column
Publer bulk-import template (`PUBLER_HEADER`); the first six columns carry
data, the rest stay empty.
"""

from __future__ import annotations

import datetime as dt
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import ClassVar, Optional

from .state import CsvRow, Platform

# ---------------------------------------------------------------------------
# Publer 12-column bulk-import template (the canonical output header)
# ---------------------------------------------------------------------------

PUBLER_HEADER = [
    "Date - Intl. format or prompt",
    "Text",
    "Link(s) - Separated by comma for FB carousels",
    "Media URL(s) - Separated by comma",
    "Title - For the video, pin, PDF ..",
    "Label(s) - Separated by comma",
    "Alt text(s) - Separated by ||",
    "Comment(s) - Separated by ||",
    "Pin board, FB album, or Google category",
    "Post subtype - I.e. story, reel, PDF ..",
    "CTA - For Facebook links or Google",
    "Reminder - For stories, reels, shorts, and TikToks",
]


def row_to_publer(row: CsvRow) -> list[str]:
    """A CsvRow rendered as a 12-field Publer template line."""
    return [
        row.date, row.text, row.link, row.media_url, row.title, row.label,
        "", "", "", "", "", "",
    ]


# ---------------------------------------------------------------------------
# Layout of the content-plan sheet (fixed, absolute spreadsheet columns)
# ---------------------------------------------------------------------------

CONTENT_SHEET = "Контент-план"


def _col(ref: str) -> int:
    """Spreadsheet column letters -> 0-based index. 'A' -> 0, 'AZ' -> 51."""
    n = 0
    for ch in ref.strip().upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


@dataclass(frozen=True)
class PlanSpec:
    """One output CSV: which platform it targets and which columns feed it.

    Column letters are ABSOLUTE positions in the content-plan sheet, so the
    layout is read straight off the spec the user maintains. The "day of week"
    column is deliberately not listed — it is derivable from the date and
    never reaches the CSV.

    MANDATORY: `date`, `time`, `media` and `text`, plus `title` for the units
    that declare one (YouTube video / shorts). "Mandatory" means a gap is
    worth a warning, not that the row is thrown away — see `slice_plan`.

    OPTIONAL: `hashtags` and `people` — a post without them is still a post.
    Both are concatenated onto the text, in the order text → hashtags →
    people, into the single Publer Text field.

    Each unit carries its OWN time column, because units that share a date
    (FB post + FB clip, IG post + IG reel, YT post + video + shorts) publish
    at different times.
    """
    suffix: str                     # file-name suffix, e.g. "FACEBOOK_OFFICIAL_POSTS"
    platform: Platform
    date: str
    time: str
    media: str
    text: str
    hashtags: Optional[str] = None
    people: Optional[str] = None    # "тэг людей", appended to the text
    title: Optional[str] = None     # YouTube video / shorts headline

    def columns(self) -> list[tuple[str, str]]:
        """(field name, column letter) for every column this unit reads."""
        declared = [
            ("date", self.date), ("time", self.time), ("text", self.text),
            ("media", self.media), ("hashtags", self.hashtags),
            ("people", self.people), ("title", self.title),
        ]
        return [(name, ref) for name, ref in declared if ref]


# The content units. One entry = one output CSV (when it has any media).
PLAN_SPECS: list[PlanSpec] = [
    PlanSpec("FACEBOOK_OFFICIAL_POSTS", Platform.FACEBOOK,
             date="A", time="C", text="E", media="G", hashtags="H"),
    PlanSpec("FACEBOOK_OFFICIAL_CLIPS", Platform.FACEBOOK,
             date="A", time="J", text="K", media="L", hashtags="M"),

    PlanSpec("INSTAGRAMM_OFFICIAL_POSTS", Platform.INSTAGRAM_BAND,
             date="O", time="Q", text="S", media="U", hashtags="V", people="W"),
    # Reels carry their own description (Z), not the post text in S.
    PlanSpec("INSTAGRAMM_OFFICIAL_REELS", Platform.INSTAGRAM_BAND,
             date="O", time="X", text="Z", media="AA", hashtags="AB", people="AC"),

    PlanSpec("INSTAGRAMM_EXCLUSIVE_POSTS", Platform.INSTAGRAM_FUNNEL,
             date="AH", time="AJ", text="AL", media="AN", hashtags="AO", people="AP"),
    PlanSpec("INSTAGRAMM_EXCLUSIVE_REELS", Platform.INSTAGRAM_FUNNEL,
             date="AH", time="AQ", text="AR", media="AS", hashtags="AT", people="AU"),

    PlanSpec("YOUTUBE_POST", Platform.YOUTUBE,
             date="AZ", time="BB", text="BD", media="BF"),
    PlanSpec("YOUTUBE_VIDEO", Platform.YOUTUBE,
             date="AZ", time="BG", title="BH", text="BI", media="BJ", hashtags="BK"),
    # Time is BL: BO is the "Shorts" marker column, per the layout.
    PlanSpec("YOUTUBE_SHORTS", Platform.YOUTUBE,
             date="AZ", time="BL", title="BM", text="BN", media="BP", hashtags="BQ"),

    PlanSpec("TIKTOK", Platform.TIKTOK,
             date="BS", time="BU", media="BX", text="BY", hashtags="BZ"),
    PlanSpec("TELEGRAMM", Platform.TELEGRAM,
             date="CB", time="CD", text="CF", media="CH"),
    PlanSpec("TWITTER", Platform.TWITTER,
             date="CJ", time="CL", text="CN", media="CP", hashtags="CQ"),
]


# ---------------------------------------------------------------------------
# The layout contract — exact header labels at exact addresses
# ---------------------------------------------------------------------------

HEADER_ROWS = 4     # labels live in sheet rows 1-4; row 5 onwards is data

# The layout, transcribed verbatim from the master template
# (FULL_FINAL_CONTENT_PLAN_HEADING.xls). EVERY labelled column is listed, not
# just the ones the converter reads: the converter addresses cells by
# position, so the cheapest way to know a sheet really is this layout is to
# check the whole header, left to right.
#
# Matched EXACTLY — only surrounding whitespace is ignored. The template mixes
# case for the same field ("Post text" vs "Post Text", "for the reels" vs
# "for the Reels"); copied as-is on purpose. Rename a column in the template
# and it must be renamed here in the same commit, or uploads are rejected.
#
# Columns absent here (N, AG, AY, BR, CA, CI, CR) are the unlabelled spacers
# between blocks and are not checked.
EXPECTED_HEADERS: dict[str, str] = {
    "A": "Date",
    "B": "Day",
    "C": "Post publication time",
    "D": "Topic",
    "E": "Post text",
    "F": "Photo/Video/Track",
    "G": "Media link for the post",
    "H": "Hashtags for the post",
    "I": "Facebook clip",
    "J": "FB clips publication time",
    "K": "FB clips description",
    "L": "Media link for the FB clip",
    "M": "Hashtags for the FB clip",
    "O": "Date",
    "P": "Day",
    "Q": "Post publication time",
    "R": "Topic",
    "S": "Post text",
    "T": "Photo/Video",
    "U": "Media link for the post",
    "V": "Hashtags for the post",
    "W": "Tags",
    "X": "Reels publication time",
    "Y": "Reels",
    "Z": "Reels Description",
    "AA": "Media link for the reels",
    "AB": "Hashtags for the reels",
    "AC": "Tags",
    "AD": "Stories",
    "AE": "Media link for the stories",
    "AF": "Hashtags for stories",
    "AH": "Date",
    "AI": "Day",
    "AJ": "Post publication time",
    "AK": "Topic",
    "AL": "Post Text",
    "AM": "Photo/Video",
    "AN": "Media link for the post",
    "AO": "Hashtags for the post",
    "AP": "Tags",
    "AQ": "Reels publication time",
    "AR": "Reels",
    "AS": "Media link for the Reels",
    "AT": "Hashtags for Reels",
    "AU": "Tags",
    "AV": "Stories",
    "AW": "Media link for the Stories",
    "AX": "Hashtags for Stories",
    "AZ": "Date",
    "BA": "Day",
    "BB": "Post publication time",
    "BC": "Новостная лента \"Записи\"",
    "BD": "Post Text",
    "BE": "Photo/Video",
    "BF": "Media link for the post",
    "BG": "Video publication time",
    "BH": "Video header",
    "BI": "Video description",
    "BJ": "Media link for the Video",
    "BK": "Hashtags for Video",
    "BL": "Shorts publication time",
    "BM": "Shorts header",
    "BN": "Shorts description",
    "BO": "Shorts",
    "BP": "Media link for the shorts",
    "BQ": "Hashtags for shorts",
    "BS": "Date",
    "BT": "Day",
    "BU": "Post publication time TikTok",
    "BV": "Topic",
    "BW": "Video",
    "BX": "Media link for the video",
    "BY": "Text/Description",
    "BZ": "Hashtags",
    "CB": "Date",
    "CC": "Day",
    "CD": "Post publication time",
    "CE": "Topic",
    "CF": "Post text",
    "CG": "Photo/Video/Track",
    "CH": "Media link for the post",
    "CJ": "Date",
    "CK": "Day",
    "CL": "Tweet publication time",
    "CM": "Topic",
    "CN": "Tweet text",
    "CO": "Photo/Video",
    "CP": "Media link for the tweet",
    "CQ": "Hashtags",
}


def checked_columns() -> list[str]:
    """Every column under contract, ordered left to right as in the sheet."""
    return sorted(EXPECTED_HEADERS, key=_col)


def units_reading(column: str) -> tuple[str, ...]:
    """Which content units break if this column is wrong. Empty for the
    columns the layout has but the converter doesn't read (Day, Topic, ...)."""
    return tuple(
        spec.suffix for spec in PLAN_SPECS
        if any(ref == column for _, ref in spec.columns())
    )


# ---------------------------------------------------------------------------
# Normalised grid — one backend per file type, same accessor interface
# ---------------------------------------------------------------------------

_MAX_COLS = 200    # cap runaway repeated/padded columns on trailing blanks
_MAX_ROWS = 1000   # content plans are ~a few hundred rows; ignore full-sheet padding


@dataclass
class _Cell:
    text: str = ""
    date: Optional[dt.date] = None
    time: Optional[tuple[int, int]] = None


def _number_cell(value: float) -> _Cell:
    """A bare number, normalised identically by every backend.

    A day fraction (0 < f < 1) is a time-of-day. A plain 0 is this plan's
    "empty" placeholder — it is neither text nor midnight.
    """
    cell = _Cell()
    if 0 < value < 1:
        total = round(value * 24 * 60) % (24 * 60)
        cell.time = (total // 60, total % 60)
    if value == 0:
        cell.text = ""
    elif float(value).is_integer():
        cell.text = str(int(value))
    else:
        cell.text = str(value)
    return cell


class _Grid:
    """A read sheet normalised to _Cell[r][c]. Backends fill `cells`.

    Merged cells need no special handling here. A merge stores its value in
    the anchor cell and leaves the rest empty, so a header merged DOWN over
    several rows still puts its text in its own column somewhere inside the
    header band — which is exactly what the band scan looks for. A header
    merged ACROSS columns is a group heading (the block titles in row 1, a
    sub-block caption) and deliberately does not count as a column's header.
    """

    def __init__(self, cells: list[list[_Cell]]):
        self.cells = cells
        self.nrows = len(cells)
        self.ncols = max((len(r) for r in cells), default=0)

    def header_label(self, r: int, c: int) -> str:
        """The label this cell contributes to column `c`."""
        return self._at(r, c).text.strip()

    def _at(self, r: int, c: Optional[int]) -> _Cell:
        if c is None or r >= self.nrows or c >= len(self.cells[r]):
            return _Cell()
        return self.cells[r][c]

    def cell_str(self, r: int, c: Optional[int]) -> str:
        return self._at(r, c).text

    def cell_date(self, r: int, c: Optional[int]) -> Optional[dt.date]:
        d = self._at(r, c).date
        return d if d and 2000 <= d.year <= 2100 else None

    def cell_time(self, r: int, c: Optional[int]) -> Optional[tuple[int, int]]:
        return self._at(r, c).time


# ---- .xls backend (xlrd) --------------------------------------------------

def _xlrd_string_record_contents(self, data):
    """xlrd's STRING-record reader, corrected for non-BMP characters.

    BIFF declares a string's length in UTF-16 code units; xlrd compares that
    against `len()` of the decoded Python string. An emoji is one Python
    character but two code units, so on any formula whose cached string result
    contains one, xlrd decides the record is unfinished and goes looking for a
    CONTINUE record that was never written:

        XLRDError: Expected CONTINUE record; found record-type 0x00BE

    Counting code units makes the comparison mean what BIFF meant. Everything
    else is xlrd's own logic, kept as-is.
    """
    from struct import unpack

    from xlrd.biffh import XLRDError, XL_CONTINUE

    bv = self.biff_version
    bk = self.book
    lenlen = (bv >= 30) + 1
    nchars_expected = unpack("<" + "BH"[lenlen - 1], data[:lenlen])[0]
    offset = lenlen
    enc = bk.encoding or bk.derive_encoding() if bv < 80 else None

    nchars_found = 0
    result = ""
    while True:
        if bv >= 80:
            enc = "utf_16_le" if data[offset] & 1 else "latin_1"
            offset += 1
        chunk = data[offset:].decode(enc)
        result += chunk
        nchars_found += len(chunk.encode("utf_16_le")) // 2
        if nchars_found == nchars_expected:
            return result
        if nchars_found > nchars_expected:
            raise XLRDError(
                f"STRING/CONTINUE: expected {nchars_expected} chars, "
                f"found {nchars_found}"
            )
        rc, _unused_len, data = bk.get_record_parts()
        if rc != XL_CONTINUE:
            raise XLRDError(f"Expected CONTINUE record; found record-type 0x{rc:04X}")
        offset = 0


@contextmanager
def _patched_xlrd():
    """Apply the fix above for the duration of one read.

    Scoped rather than installed at import: this reaches into another
    library's internals, so it should be in force only where we need it.
    """
    from xlrd import sheet as xlrd_sheet

    original = xlrd_sheet.Sheet.string_record_contents
    xlrd_sheet.Sheet.string_record_contents = _xlrd_string_record_contents
    try:
        yield
    finally:
        xlrd_sheet.Sheet.string_record_contents = original


def _read_xls(path: Path) -> _Grid:
    import xlrd

    with _patched_xlrd():
        wb = xlrd.open_workbook(str(path))
    try:
        sh = wb.sheet_by_name(CONTENT_SHEET)
    except xlrd.XLRDError:
        sh = wb.sheet_by_index(0)

    def to_cell(r: int, c: int) -> _Cell:
        ctype = sh.cell_type(r, c)
        val = sh.cell_value(r, c)
        cell = _Cell()
        if ctype == xlrd.XL_CELL_EMPTY:
            return cell
        if ctype == xlrd.XL_CELL_DATE:
            d = xlrd.xldate_as_datetime(float(val), wb.datemode)
            cell.date = d.date()
            cell.time = (d.hour, d.minute)
            cell.text = f"{d:%Y/%m/%d %H:%M}"
            return cell
        if ctype == xlrd.XL_CELL_NUMBER:
            return _number_cell(float(val))
        # string / bool
        s = str(val).strip()
        cell.text = "" if s in ("0", "0.0") else s
        return cell

    cells = [[to_cell(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    return _Grid(cells)


# ---- .xlsx backend (openpyxl) ---------------------------------------------

def _read_xlsx(path: Path) -> _Grid:
    from openpyxl import load_workbook

    # data_only: read the cached results of formulas, not the formulas.
    wb = load_workbook(str(path), data_only=True, read_only=True)
    sheet = wb[CONTENT_SHEET] if CONTENT_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]

    def to_cell(val) -> _Cell:
        if val is None:                       # includes merged-range followers
            return _Cell()
        # datetime before date, bool before int — both are subclasses.
        if isinstance(val, dt.datetime):
            return _Cell(f"{val:%Y/%m/%d %H:%M}", val.date(), (val.hour, val.minute))
        if isinstance(val, dt.date):
            return _Cell(f"{val:%Y/%m/%d}", val)
        if isinstance(val, dt.time):
            return _Cell(f"{val:%H:%M}", None, (val.hour, val.minute))
        if isinstance(val, dt.timedelta):     # a duration-formatted time cell
            minutes = int(val.total_seconds()) // 60 % (24 * 60)
            h, m = divmod(minutes, 60)
            return _Cell(f"{h:02d}:{m:02d}", None, (h, m))
        if isinstance(val, bool):
            return _Cell(str(val))
        if isinstance(val, (int, float)):
            return _number_cell(float(val))
        s = str(val).strip()
        return _Cell("" if s in ("0", "0.0") else s)

    try:
        cells = [
            [to_cell(v) for v in row]
            for row in sheet.iter_rows(max_row=_MAX_ROWS, max_col=_MAX_COLS, values_only=True)
        ]
    finally:
        # read_only mode keeps the file handle open, which on Windows blocks
        # the caller from deleting its temp copy of the upload.
        wb.close()
    return _Grid(cells)


# ---- .ods backend (odfpy) -------------------------------------------------

def _read_ods(path: Path) -> _Grid:
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = load(str(path))
    tables = doc.spreadsheet.getElementsByType(Table)
    sheet = None
    for t in tables:
        if t.getAttribute("name") == CONTENT_SHEET:
            sheet = t
            break
    if sheet is None and tables:
        sheet = tables[0]
    if sheet is None:
        return _Grid([])

    def cell_text(cell) -> str:
        parts = []
        for p in cell.getElementsByType(P):
            parts.append(str(p))
        return "\n".join(parts).strip()

    def parse_cell(cell) -> _Cell:
        vtype = cell.getAttribute("valuetype")
        out = _Cell()
        if vtype == "date":
            raw = cell.getAttribute("datevalue") or ""
            try:
                if "T" in raw:
                    d = dt.datetime.fromisoformat(raw)
                    out.date = d.date()
                    out.time = (d.hour, d.minute)
                    out.text = f"{d:%Y/%m/%d %H:%M}"
                else:
                    out.date = dt.date.fromisoformat(raw[:10])
                    out.text = f"{out.date:%Y/%m/%d}"
            except ValueError:
                out.text = cell_text(cell)
        elif vtype == "time":
            raw = cell.getAttribute("timevalue") or ""   # e.g. PT18H34M00S
            h = m = 0
            num = ""
            for ch in raw:
                if ch.isdigit():
                    num += ch
                elif ch == "H":
                    h = int(num or 0); num = ""
                elif ch == "M":
                    m = int(num or 0); num = ""
                elif ch in ("S", "T", "P"):
                    num = ""
            out.time = (h, m)
            out.text = f"{h:02d}:{m:02d}"
        elif vtype == "float":
            out = _number_cell(float(cell.getAttribute("value") or "0"))
        else:
            out.text = cell_text(cell)
        return out

    cells: list[list[_Cell]] = []
    for row in sheet.getElementsByType(TableRow):
        row_rep = int(row.getAttribute("numberrowsrepeated") or 1)
        built: list[_Cell] = []
        # Iterate children in document order so merged cells stay aligned:
        # a spanned cell is one <table-cell> followed by N-1
        # <covered-table-cell> siblings that still occupy real columns.
        for child in row.childNodes:
            qname = getattr(child, "qname", None)
            if qname is None:
                continue
            local = qname[1]
            if local not in ("table-cell", "covered-table-cell"):
                continue
            rep = int(child.getAttribute("numbercolumnsrepeated") or 1)
            parsed = _Cell() if local == "covered-table-cell" else parse_cell(child)
            # Don't materialise huge trailing-blank repeats.
            if rep > _MAX_COLS and not parsed.text:
                rep = 1
            built.extend(_Cell(parsed.text, parsed.date, parsed.time) for _ in range(rep))
            if len(built) >= _MAX_COLS:
                break
        # Empty repeated rows are just blanks; keep at most a couple.
        row_rep = 1 if row_rep > 4 else row_rep
        for _ in range(row_rep):
            cells.append(list(built))
            if len(cells) >= _MAX_ROWS:   # ignore full-sheet trailing padding
                return _Grid(cells)
    return _Grid(cells)


PLAN_EXTENSIONS = (".ods", ".xls", ".xlsx")


def _open(path: Path) -> _Grid:
    suffix = path.suffix.lower()
    if suffix == ".ods":
        return _read_ods(path)
    if suffix == ".xls":
        return _read_xls(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(
        f"Unsupported content-plan format: {suffix} "
        f"(use {', '.join(PLAN_EXTENSIONS)})"
    )


# ---------------------------------------------------------------------------
# Slicing
# ---------------------------------------------------------------------------

@dataclass
class PlanCsv:
    """One content unit's output file sliced from the plan."""
    name: str            # e.g. "AUG2026_PLAN_FACEBOOK_OFFICIAL_POSTS" (no extension)
    platform: Platform
    unit: str = ""       # the PlanSpec suffix this came from, e.g. "TIKTOK"
    rows: list[CsvRow] = field(default_factory=list)
    # Sheet rows that made it into the file with gaps still in them. Carried
    # on the file itself so a half-finished CSV can say so where it is
    # downloaded — a warning elsewhere on the page is easy to import past.
    unfinished: tuple[int, ...] = ()


@dataclass(frozen=True)
class HeaderMismatch:
    """A column whose header isn't the label the layout puts there."""
    column: str                 # spreadsheet column letter, e.g. "BL"
    expected: str               # the label the master template shows there
    found: str                  # what the sheet shows instead ("" if nothing)
    units: tuple[str, ...]      # the content units this column feeds

    def message(self) -> str:
        if not self.found:
            return f"Колонка {self.column}: отсутствует заголовок «{self.expected}»"
        return (f"Колонка {self.column}: ожидается «{self.expected}», "
                f"найдено «{self.found}»")


class PlanLayoutError(ValueError):
    """The plan's layout is wrong — raised before a single row is converted.

    The converter addresses columns by position, so a shifted or truncated
    sheet would not fail loudly: it would quietly slice the wrong cells, or
    nothing at all. This turns that into a refusal to process the file.
    """

    def __init__(self, mismatches: list[HeaderMismatch]):
        self.mismatches = mismatches
        super().__init__(
            "; ".join(m.message() for m in mismatches) + " — необходимо поправить макет"
        )


def check_layout(grid: _Grid) -> list[HeaderMismatch]:
    """Walk the header left to right; report every column that isn't the
    layout's, exactly.

    The label must appear at its own column somewhere in the header band
    (rows 1-4) — the band is scanned rather than a fixed row because the
    template puts the YouTube post sub-headers a row below the rest. Matching
    is exact apart from surrounding whitespace.
    """
    band = range(min(HEADER_ROWS, grid.nrows))
    out: list[HeaderMismatch] = []
    for column in checked_columns():
        index = _col(column)
        labels = [grid.header_label(r, index) for r in band]
        if EXPECTED_HEADERS[column] in labels:
            continue
        out.append(HeaderMismatch(
            column, EXPECTED_HEADERS[column], _shown_label(labels), units_reading(column)
        ))
    return out


def _shown_label(labels: list[str]) -> str:
    """What to quote back as "found".

    Row 1 holds the block titles (FACEBOOK, YOUTUBE) merged over whole blocks
    and row 4 holds column numbering; neither is this column's header, so
    prefer a real label from the rows in between.
    """
    for label in labels[1:]:
        if label and not label.replace(".", "").isdigit():
            return label
    return next((label for label in labels if label), "")


@dataclass(frozen=True)
class PlanWarning:
    """Something about one plan row worth telling the author.

    Advisory only — nothing here stops a file from being written. The two
    kinds below differ in what they cost: an unfinished row reaches Publer
    with holes in it, while a post with no hashtags is complete but bare.
    """
    kind: ClassVar[str] = "warning"

    unit: str                   # the unit's file suffix
    sheet_row: int              # 1-based row number, as shown in the spreadsheet
    columns: tuple[str, ...]    # the cells at issue, left to right

    def labelled(self) -> list[str]:
        """The cells named as the sheet names them: «Post text» (E)."""
        return [f"«{EXPECTED_HEADERS.get(c, c)}» ({c})" for c in self.columns]

    def message(self) -> str:
        raise NotImplementedError


@dataclass(frozen=True)
class IncompleteRow(PlanWarning):
    """Some mandatory fields are filled and others aren't.

    The row is still converted — the gap may be deliberate, to be filled in
    inside Publer — so this never blocks anything. It exists because the other
    case, someone who got interrupted, looks exactly the same from here and is
    far more common.
    """
    kind: ClassVar[str] = "incomplete"

    def message(self) -> str:
        one = len(self.columns) == 1
        return (f"{self.unit}, строка {self.sheet_row}: похоже, "
                f"{'не заполнено поле' if one else 'не заполнены поля'} "
                f"{', '.join(self.labelled())}. Пожалуйста, проверьте.")


@dataclass(frozen=True)
class MissingHashtags(PlanWarning):
    """A complete post whose hashtag cell is empty.

    Hashtags are optional — the row converts and publishes either way — so
    this never blocks anything. It is here because leaving them out is far
    more often an oversight than a decision.
    """
    kind: ClassVar[str] = "no_hashtags"

    def message(self) -> str:
        return (f"{self.unit}, строка {self.sheet_row}: не заполнено поле "
                f"{', '.join(self.labelled())} — пост уйдёт без хэштегов. "
                f"Пожалуйста, проверьте.")


@dataclass(frozen=True)
class MalformedHashtags(PlanWarning):
    """A hashtag cell with content but not a single `#` in it.

    `arcticdreams metal` instead of `#arcticdreams #metal` publishes as
    plain trailing words — technically fine, almost certainly a typo.
    Purely mechanical: the check is "does the cell contain `#` at all",
    never "are these the right hashtags".
    """
    kind: ClassVar[str] = "malformed_hashtags"

    def message(self) -> str:
        return (f"{self.unit}, строка {self.sheet_row}: в поле "
                f"{', '.join(self.labelled())} есть текст, но нет ни одного "
                f"«#» — теги уйдут как обычные слова. Пожалуйста, проверьте.")


@dataclass
class PlanSlices:
    """Everything one pass over a content plan produced."""
    files: list[PlanCsv] = field(default_factory=list)
    warnings: list[PlanWarning] = field(default_factory=list)


def _safe_name(label: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in label)


def has_media(cell: str) -> bool:
    """Does this cell hold a media link at all?

    "Filled in" is stricter for the media column than for the text ones: an
    http(s) URL must actually be there. Whether that URL is a *temporary*
    Publer upload is the linter's `media_url_permanent` rule to judge, not
    ingestion's.
    """
    for token in cell.replace(",", " ").split():
        low = token.lower()
        if low.startswith("http://") or low.startswith("https://"):
            return True
    return False


def _combine_text(*parts: str) -> str:
    """Text, hashtags and people tags become one Publer Text field."""
    return "\n\n".join(p.strip() for p in parts if p and p.strip())


def _build_date(d: Optional[dt.date], hm: Optional[tuple[int, int]]) -> str:
    """Whatever of the date is filled in — nothing is invented.

    A row can reach the CSV with no time on purpose: the author may mean to
    schedule it inside Publer. Half a date is what the plan says, so half a
    date is what gets written.
    """
    if d is None:
        return ""
    if hm is None:
        return f"{d:%Y/%m/%d}"
    return f"{d:%Y/%m/%d} {hm[0]:02d}:{hm[1]:02d}"


def slice_plan(path: Path, source_name: Optional[str] = None) -> PlanSlices:
    """Read a content-plan spreadsheet into one CSV per content unit.

    Raises `PlanLayoutError` before converting anything if the sheet's layout
    isn't the template's.

    A row is a content unit only when **every** mandatory field of that unit is
    filled — date, time, media, text, and the title where the unit has one.
    A row with no content of its own — no text, media or title — means the
    unit wasn't planned that day and is passed over quietly; the date and time
    columns don't count, both being scaffolding filled in a month ahead.
    Everything else is converted with whatever it holds, and its gaps come
    back in `.warnings`, as does a finished post with no hashtags.

    `source_name` is the plan's original file name — it prefixes every output
    file. Pass it explicitly when `path` is a temp copy of an upload.
    """
    path = Path(path)
    grid = _open(path)

    mismatches = check_layout(grid)
    if mismatches:
        raise PlanLayoutError(mismatches)

    stem = _safe_name(Path(source_name or path.name).stem)

    def get(r: int, ref: Optional[str]) -> str:
        """Cell text by column letter; an unset column reads as empty."""
        return grid.cell_str(r, _col(ref)) if ref else ""

    out = PlanSlices()
    for spec in PLAN_SPECS:
        date_col = _col(spec.date)
        time_col = _col(spec.time)

        rows: list[CsvRow] = []
        unfinished: list[int] = []
        # Start below the header band: its labels sit in the text columns, so
        # scanning it would report every header row as a half-filled post.
        for r in range(HEADER_ROWS, grid.nrows):
            date = grid.cell_date(r, date_col)
            time = grid.cell_time(r, time_col)   # this unit's own time column
            media = get(r, spec.media).strip()
            text = get(r, spec.text).strip()
            title = get(r, spec.title).strip()
            hashtags = get(r, spec.hashtags).strip()

            # Hashtags and people tags are the only optional fields; the rest
            # are all needed before the row can be a post.
            mandatory = [
                (spec.date, date is not None),
                (spec.time, time is not None),
                (spec.text, bool(text)),
                (spec.media, has_media(media)),
            ]
            if spec.title:
                mandatory.append((spec.title, bool(title)))
            mandatory.sort(key=lambda pair: _col(pair[0]))
            missing = tuple(column for column, filled in mandatory if not filled)

            # Has this unit been started at all? Date and time don't count:
            # both are scaffolding laid down a month ahead — the date shared by
            # the whole block, the time a formula dragged down the column — so
            # neither says anything about this unit. A row carrying only those
            # is a day nobody has planned, and it is passed over in silence.
            started = any(filled for column, filled in mandatory
                          if column not in (spec.date, spec.time))
            if not started:
                continue

            # Whatever is filled goes in, gaps and all. A missing time may be
            # deliberate — scheduled later inside Publer — so ingestion doesn't
            # get to decide the row isn't wanted.
            rows.append(CsvRow(
                row_index=len(rows),
                platform=spec.platform,
                date=_build_date(date, time),
                text=_combine_text(text, hashtags, get(r, spec.people)),
                link="",
                media_url=media,
                title=title,
                label=spec.suffix,
            ))

            if missing:
                out.warnings.append(IncompleteRow(spec.suffix, r + 1, missing))
                unfinished.append(r + 1)
            elif spec.hashtags and not hashtags:
                # Nothing wrong with the post; it just has no hashtags.
                out.warnings.append(
                    MissingHashtags(spec.suffix, r + 1, (spec.hashtags,))
                )
            if spec.hashtags and hashtags and "#" not in hashtags:
                out.warnings.append(
                    MalformedHashtags(spec.suffix, r + 1, (spec.hashtags,))
                )

        # Not one started row -> the unit wasn't planned, so no file.
        if rows:
            out.files.append(PlanCsv(
                name=f"{stem}_{spec.suffix}",
                platform=spec.platform,
                unit=spec.suffix,
                rows=rows,
                unfinished=tuple(unfinished),
            ))

    return out
